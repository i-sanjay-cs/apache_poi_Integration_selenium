from openpyxl import load_workbook

class ExcelDataProvider:
    @staticmethod
    def get_test_data(file_path, sheet_name):
        workbook = load_workbook(file_path)
        sheet = workbook[sheet_name]
        data = []

        for row in sheet.iter_rows(min_row=2, values_only=True):
            data.append(row)

        workbook.close()
        return data
